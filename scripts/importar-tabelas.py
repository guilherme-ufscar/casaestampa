"""
Script de importação das tabelas de preço para o banco Supabase.
- Desativa tecidos/trilhos antigos (preserva FK dos orçamentos existentes)
- Importa tecidos de cortina, reajuste, cinzas e decoração da Deccor
- Importa trilhos Venetillo e Rio Flex
- Atualiza configurações de confecção e instalação
"""

import re
import os
import glob
import psycopg2
from cuid2 import cuid_wrapper
import openpyxl

cuid = cuid_wrapper()
DB_URL = 'postgresql://postgres.eqouuvhwbcqzerbywvos:CoderMaster20262026@aws-1-sa-east-1.pooler.supabase.com:5432/postgres'

# Encontrar arquivos por glob para evitar problemas de encoding
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
deccor_files = glob.glob('Tabela*DECCOR*.xlsx')
DECCOR_FILE = deccor_files[0] if deccor_files else None
TRILHO_FILE = 'TABELA DE TRILHO.xlsx'


def parse_largura(valor):
    """Parse '3,00 m' ou '2,80/3,00m' -> float (maior valor)"""
    if not valor:
        return 3.00
    valor = str(valor).strip().lower().replace('m', '').strip()
    # Caso "2,80/3,00"
    if '/' in valor:
        partes = valor.split('/')
        nums = []
        for p in partes:
            p = p.strip().replace(',', '.')
            try:
                nums.append(float(p))
            except ValueError:
                pass
        return max(nums) if nums else 3.00
    valor = valor.replace(',', '.')
    try:
        return float(valor)
    except ValueError:
        return 3.00


def parse_preco(valor):
    """Parse preço: retorna float ou None se inválido"""
    if valor is None:
        return None
    valor = str(valor).strip()
    if valor in ('-', '', 'None', '-\n'):
        return None
    valor = valor.replace(',', '.')
    try:
        v = float(valor)
        return v if v > 0 else None
    except ValueError:
        return None


def importar_tecidos_cortina(ws):
    """Sheet 'Tecidos Cortina': colunas Código, Produto, +/-, Largura, Metro"""
    tecidos = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        # Pular cabeçalhos e linhas informativas
        if not row or len(row) < 5:
            continue
        codigo, produto, _, largura, metro = row[0], row[1], row[2], row[3], row[4]

        # Precisa ter produto e preço válido
        if not produto or not isinstance(produto, str):
            continue
        if str(produto).strip().upper() in ('PRODUTO', ''):
            continue

        preco = parse_preco(metro)
        if preco is None:
            continue

        larg = parse_largura(largura)
        nome = produto.strip()

        # Adicionar código ao nome para identificação
        if codigo and str(codigo).strip():
            nome_final = f"{str(codigo).strip()} - {nome}"
        else:
            nome_final = nome

        tecidos.append({
            'id': cuid(),
            'nome': nome_final,
            'larguraMaxima': larg,
            'valorMetro': preco,
            'tipo': 'PRINCIPAL',
            'ativo': True,
        })
    return tecidos


def importar_tecidos_reajuste(ws):
    """Sheet 'REAJUSTE 1505': colunas Código, Produto, +/-, Largura, Atacado, Varejo"""
    tecidos = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 6:
            continue
        codigo, produto, _, largura, atacado, varejo = row[0], row[1], row[2], row[3], row[4], row[5]

        if not produto or not isinstance(produto, str):
            continue
        if str(produto).strip().upper() in ('PRODUTO', 'CÓDIGO', ''):
            continue
        if str(codigo).strip().upper() == 'CÓDIGO':
            continue

        preco = parse_preco(varejo)
        if preco is None:
            continue

        larg = parse_largura(largura)
        nome = produto.strip()

        if codigo and str(codigo).strip():
            nome_final = f"{str(codigo).strip()} - {nome}"
        else:
            nome_final = nome

        tecidos.append({
            'id': cuid(),
            'nome': nome_final,
            'larguraMaxima': larg,
            'valorMetro': preco,
            'tipo': 'PRINCIPAL',
            'ativo': True,
        })
    return tecidos


def importar_tecidos_cinzas(ws):
    """Sheet 'Cinzas': mesma estrutura do REAJUSTE"""
    tecidos = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 6:
            continue
        codigo, produto, _, largura, atacado, varejo = row[0], row[1], row[2], row[3], row[4], row[5]

        if not produto or not isinstance(produto, str):
            continue
        if str(produto).strip().upper() in ('PRODUTO', 'CÓDIGO', ''):
            continue
        if str(codigo).strip().upper() == 'CÓDIGO':
            continue

        preco = parse_preco(varejo)
        if preco is None:
            continue

        larg = parse_largura(largura)
        nome = produto.strip()

        if codigo and str(codigo).strip():
            nome_final = f"{str(codigo).strip()} - {nome}"
        else:
            nome_final = nome

        tecidos.append({
            'id': cuid(),
            'nome': nome_final,
            'larguraMaxima': larg,
            'valorMetro': preco,
            'tipo': 'PRINCIPAL',
            'ativo': True,
        })
    return tecidos


def importar_tecidos_decoracao(ws):
    """Sheet 'Tecidos Decoração': colunas Código, Produto, Largura, +/-, Atacado, Varejo"""
    tecidos = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 6:
            continue
        codigo, produto, largura, _, atacado, varejo = row[0], row[1], row[2], row[3], row[4], row[5]

        if not produto or not isinstance(produto, str):
            continue
        if str(produto).strip().upper() in ('PRODUTO', 'CÓDIGO', ''):
            continue
        if str(codigo).strip().upper() in ('CÓDIGO', ''):
            continue

        preco = parse_preco(varejo)
        if preco is None:
            preco = parse_preco(atacado)
        if preco is None:
            continue

        larg = parse_largura(largura)
        nome = produto.strip()

        if codigo and str(codigo).strip():
            nome_final = f"{str(codigo).strip()} - {nome}"
        else:
            nome_final = nome

        tecidos.append({
            'id': cuid(),
            'nome': nome_final,
            'larguraMaxima': larg,
            'valorMetro': preco,
            'tipo': 'DECORACAO',
            'ativo': True,
        })
    return tecidos


def importar_trilhos(ws):
    """Sheet 'TRILHO': Venetillo e Rio Flex"""
    trilhos = []
    fornecedor = None

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        cell0 = str(row[0]).strip() if row[0] else ''

        # Detectar fornecedor
        if 'VENETILLO' in cell0.upper():
            fornecedor = 'Venetillo'
            continue
        elif 'RIO FLEX' in cell0.upper():
            fornecedor = 'Rio Flex'
            continue

        if not fornecedor:
            continue

        # Pular cabeçalhos
        if not cell0 or cell0.upper() in ('', 'TABELA TRILHO CORTINA'):
            continue

        # Pegar valor
        valor = parse_preco(row[1]) if len(row) > 1 else None
        if valor is None:
            continue

        nome = f"{cell0} ({fornecedor})"

        trilhos.append({
            'id': cuid(),
            'nome': nome,
            'valorUnitario': valor,
            'ativo': True,
        })

    return trilhos


def importar_confeccao_instalacao(wb):
    """Sheets CONFECÇÃO e INSTALAÇÃO -> ConfiguracaoCalculo"""
    configs = {}

    # Sheet Confecção
    ws = wb['CONFECÇÃO']
    mapa_confeccao = {
        'PREGA MACHO': 'confeccao_prega_macho',
        'PREGA FEMEA': 'confeccao_prega_femea',
        'PREGA RETA FORRO': 'confeccao_prega_reta',
        'PREGA FRANZIDA': 'confeccao_prega_franzida',
        'PREGA AMERICANA': 'confeccao_prega_americana',
        'PREGA WAVE': 'confeccao_wave',
        'PREGA SOFT WAVE': 'confeccao_soft_wave',
        'CORTINA DE VARÃO ILHOES': 'confeccao_varao',
        'CORTINA DE VARÃO ILHOES': 'confeccao_varao',
    }

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        nome = str(row[0]).strip().upper()
        for chave_planilha, chave_config in mapa_confeccao.items():
            if chave_planilha in nome:
                valor = parse_preco(row[1])
                if valor:
                    configs[chave_config] = str(valor)
                break

    # Sheet Instalação
    ws = wb['INSTALAÇÃO']
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        nome = str(row[0]).strip().upper()
        if 'CORTINA' in nome:
            valor = parse_preco(row[1])
            if valor:
                configs['instalacao_valor_m2'] = str(valor)
                break

    return configs


def main():
    print("=" * 60)
    print("IMPORTAÇÃO DE TABELAS DE PREÇO - CASA ESTAMPA")
    print("=" * 60)

    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    # 1. Desativar tecidos e trilhos antigos
    print("\n[1/6] Desativando tecidos e trilhos antigos...")
    cur.execute('UPDATE "Tecido" SET ativo = false')
    cur.execute('UPDATE "TrilhoVarao" SET ativo = false')
    conn.commit()
    print("  OK - Tecidos e trilhos antigos desativados")

    # 2. Importar tecidos da Deccor
    print("\n[2/6] Lendo planilha Deccor...")
    wb_deccor = openpyxl.load_workbook(DECCOR_FILE, data_only=True)

    tecidos_cortina = importar_tecidos_cortina(wb_deccor['Tecidos Cortina'])
    print(f"  Tecidos Cortina: {len(tecidos_cortina)} itens")

    tecidos_reajuste = importar_tecidos_reajuste(wb_deccor['REAJUSTE 1505'])
    print(f"  Tecidos Reajuste: {len(tecidos_reajuste)} itens")

    tecidos_cinzas = importar_tecidos_cinzas(wb_deccor['Cinzas'])
    print(f"  Tecidos Cinzas: {len(tecidos_cinzas)} itens")

    tecidos_decoracao = importar_tecidos_decoracao(wb_deccor['Tecidos Decoração'])
    print(f"  Tecidos Decoração: {len(tecidos_decoracao)} itens")

    todos_tecidos = tecidos_cortina + tecidos_reajuste + tecidos_cinzas + tecidos_decoracao
    print(f"  TOTAL tecidos: {len(todos_tecidos)}")

    # 3. Inserir tecidos
    print("\n[3/6] Inserindo tecidos no banco...")
    insert_sql = '''INSERT INTO "Tecido" (id, nome, "larguraMaxima", "valorMetro", tipo, ativo)
                    VALUES (%s, %s, %s, %s, %s, %s)'''

    batch = [(t['id'], t['nome'], t['larguraMaxima'], t['valorMetro'], t['tipo'], t['ativo']) for t in todos_tecidos]
    cur.executemany(insert_sql, batch)
    conn.commit()
    print(f"  OK - {len(batch)} tecidos inseridos")

    # 4. Importar trilhos
    print("\n[4/6] Lendo planilha de trilhos...")
    wb_trilho = openpyxl.load_workbook(TRILHO_FILE, data_only=True)
    trilhos = importar_trilhos(wb_trilho['TRILHO'])
    print(f"  Trilhos encontrados: {len(trilhos)} itens")

    # 5. Inserir trilhos
    print("\n[5/6] Inserindo trilhos no banco...")
    insert_trilho_sql = '''INSERT INTO "TrilhoVarao" (id, nome, "valorUnitario", ativo)
                           VALUES (%s, %s, %s, %s)'''

    batch_trilhos = [(t['id'], t['nome'], t['valorUnitario'], t['ativo']) for t in trilhos]
    cur.executemany(insert_trilho_sql, batch_trilhos)
    conn.commit()
    print(f"  OK - {len(batch_trilhos)} trilhos inseridos")

    # 6. Atualizar configurações de confecção/instalação
    print("\n[6/6] Atualizando configurações de confecção e instalação...")
    configs = importar_confeccao_instalacao(wb_trilho)

    for chave, valor in configs.items():
        cur.execute('''
            INSERT INTO "ConfiguracaoCalculo" (id, chave, valor) VALUES (%s, %s, %s)
            ON CONFLICT (chave) DO UPDATE SET valor = EXCLUDED.valor
        ''', (cuid(), chave, valor))

    conn.commit()
    print(f"  OK - {len(configs)} configs atualizadas: {list(configs.keys())}")

    # Resumo final
    print("\n" + "=" * 60)
    print("RESUMO FINAL")
    print("=" * 60)
    cur.execute('SELECT tipo, count(*) FROM "Tecido" WHERE ativo = true GROUP BY tipo ORDER BY tipo')
    for row in cur.fetchall():
        print(f"  Tecidos {row[0]}: {row[1]}")
    cur.execute('SELECT count(*) FROM "TrilhoVarao" WHERE ativo = true')
    print(f"  Trilhos ativos: {cur.fetchone()[0]}")
    cur.execute('SELECT count(*) FROM "Tecido" WHERE ativo = false')
    print(f"  Tecidos desativados (antigos): {cur.fetchone()[0]}")
    cur.execute('SELECT count(*) FROM "TrilhoVarao" WHERE ativo = false')
    print(f"  Trilhos desativados (antigos): {cur.fetchone()[0]}")

    conn.close()
    print("\nOK - Importacao concluida com sucesso!")


if __name__ == '__main__':
    main()
